import re

# 이메일 형식을 추출하는 코드 만들기
test_string = '''
aaa@bbb.com
123@abc.co.kr
test@hello.kr
ok@ok.co.kr
ok@ok.co.kr
ok@ok.co.kr
no.co.kr
'''

results = re.findall(r'[\w\.-]+@[\w\.-]+', test_string)
print(results)

# 리스트에서 중복 내용 제거하는 코드 만들기
results = list(set(results))
print(results)

# 사이트에서 이메일 수집하는 코드 만들기
import requests

url = 'https://sports.news.naver.com/news?oid=117&aid=0003648656'
headers = {
    'User-Agent': 'Mozilla/5.0',
    'Content-Type':'text/html; charset=utf-8'
}

response = requests.get(url,headers=headers)

results = re.findall(r'[\w\.-]+@[\w\.-]+', response.text)

results = list(set(results))
print(results)


# 수집한 이메일 주소를 엑셀에 저장하는 코드 만들기
from openpyxl import load_workbook
from openpyxl import Workbook

try:
    wb = load_workbook(r'email.xlsx', data_only=True)
    sheet = wb.active

except:
    wb = Workbook()
    sheet = wb.active

for result in results:
    sheet.append([result])

wb.save(r'email.xlsx')
